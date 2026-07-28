import hashlib
from io import BytesIO

import pandas as pd
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

PALETTE = [
    {"bg": "#e7f1ff", "fg": "#0d6efd"},
    {"bg": "#f3e8ff", "fg": "#9333ea"},
    {"bg": "#d1e7dd", "fg": "#0f5132"},
    {"bg": "#fff3cd", "fg": "#664d03"},
    {"bg": "#f8d7da", "fg": "#842029"},
    {"bg": "#cff4fc", "fg": "#055160"},
    {"bg": "#ffe5d0", "fg": "#a15c07"},
    {"bg": "#e2e3ff", "fg": "#4338ca"},
    {"bg": "#fce7f3", "fg": "#9d174d"},
    {"bg": "#e7e9eb", "fg": "#41464b"},
]

def color_for(value):
    idx = int(hashlib.md5(str(value).encode()).hexdigest(), 16) % len(PALETTE)
    return PALETTE[idx]

def badge_html(value, css_class="badge-soft"):
    if not value:
        return ""
    c = color_for(value)
    return f'<span class="{css_class}" style="background:{c["bg"]};color:{c["fg"]};">{value}</span>'

ASESOR_CORTO = {
    'HUGO ENRIQUE PÉREZ RAMÍREZ':     'HUGO PÉREZ',
    'JOSÉ ALVARO MARTÍNEZ ESPEJEL':   'ALVARO MARTÍNEZ',
    'MAURICIO GUTIÉRREZ PÉREZ PALMA': 'MAURICIO GUTIÉRREZ',
}

def dataframe_to_excel(df, sheet_name, currency_cols=None):
    """Exporta un DataFrame a Excel con encabezados en negrita, filtros
    activos (autofiltro) y formato de moneda en las columnas indicadas."""
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]

        for cell in ws[1]:
            cell.font = Font(bold=True)

        ws.auto_filter.ref = ws.dimensions

        for col in (currency_cols or []):
            if col not in df.columns:
                continue
            letter = get_column_letter(df.columns.get_loc(col) + 1)
            for row in range(2, ws.max_row + 1):
                ws[f"{letter}{row}"].number_format = '$#,##0.00'

    return output.getvalue()

def avatar_html(name):
    if not name:
        return ""
    parts = str(name).split()
    initials = "".join(p[0] for p in parts[:2]).upper()
    c = color_for(name)
    return (
        '<span style="display:inline-flex;align-items:center;gap:8px;">'
        f'<span style="display:inline-flex;align-items:center;justify-content:center;'
        f'width:24px;height:24px;border-radius:50%;background:{c["bg"]};color:{c["fg"]};'
        f'font-size:10.5px;font-weight:700;flex-shrink:0;">{initials}</span>'
        f'<span>{name}</span></span>'
    )
